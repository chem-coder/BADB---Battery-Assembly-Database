#!/usr/bin/env python3
"""Verify our SOH formula == colleague's SOH column.

For each summary sheet, walk blocks in order. A DChg.Cap block is followed by
its SOH block (same rate-step, same cells). Compute SOH = DChg(n)/DChg(first
valid)×100 from the capacity block and compare, cell by cell, cycle by cycle,
against the SOH block. Report the max absolute difference per sheet.
"""
import openpyxl
import import_ref_cycling as M


def parse_all_blocks(ws):
    """Ordered list of (kind, rate, {cell: [(cycle, val)]}) for kind in dchg/soh."""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    header_rows = [i for i, row in enumerate(rows) if any(M._is_header_cell(c) for c in row)]
    starts = []  # (kind, rate, hi, cyc_col)
    for i, row in enumerate(rows):
        for j, c in enumerate(row):
            if not isinstance(c, str):
                continue
            kind = None
            if 'Cap' in c and c.strip().startswith('D'):
                kind = 'dchg'
            elif 'SOH' in c:
                kind = 'soh'
            if kind:
                rate = None
                for up in (i - 1, i - 2):
                    if up >= 0:
                        v = rows[up][j] if j < len(rows[up]) else None
                        if isinstance(v, str) and v.strip() and not M._is_header_cell(v):
                            rate = v.strip(); break
                starts.append((kind, rate, i, j))
                break
    out = []
    for kind, rate, hi, cyc_c in starts:
        hdr = rows[hi]
        end = min([h for h in header_rows if h > hi], default=len(rows))
        data = rows[hi + 1:end]
        cycles = [int(r[cyc_c]) if cyc_c < len(r) and isinstance(r[cyc_c], (int, float)) else None for r in data]
        cells = {}
        for j in range(cyc_c + 1, ws.max_column):
            label = hdr[j] if j < len(hdr) else None
            if label is None:
                break
            lab = str(label).strip()
            series = []
            for k, r in enumerate(data):
                v = r[j] if j < len(r) else None
                if cycles[k] is not None and isinstance(v, (int, float)):
                    series.append((cycles[k], float(v)))
            if series:
                cells[lab] = series
        if cells:
            out.append((kind, rate or '?', cells))
    return out


def soh_from_cap(series):
    base = next((v for _, v in series if v and v > 0), None)
    if not base:
        return {}
    return {c: (v / base) * 100 for c, v in series if v and v > 0}


def horizontal_pairs(ws):
    """Layout where SOH columns sit beside capacity columns in the same block.
    Classify each column by its first value (cap ~0.05–12, SOH ~50–200), then
    pair the i-th cap column with the i-th SOH column. Returns list of
    (cap_series, soh_series)."""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    header_rows = [i for i, row in enumerate(rows) if any(M._is_header_cell(c) for c in row)]
    out = []
    for i, row in enumerate(rows):
        cyc_c = next((j for j, c in enumerate(row) if M._is_dchg_header(c)), None)
        if cyc_c is None:
            continue
        end = min([h for h in header_rows if h > i], default=len(rows))
        data = rows[i + 1:end]
        cycles = [int(r[cyc_c]) if cyc_c < len(r) and isinstance(r[cyc_c], (int, float)) else None for r in data]
        hdr = rows[i]
        caps, sohs = [], []
        for j in range(cyc_c + 1, ws.max_column):
            label = hdr[j] if j < len(hdr) else None
            if label is None:
                break
            if str(label).strip() == 'SOH':  # bare axis marker
                continue
            ser = [(cycles[k], float(r[j])) for k, r in enumerate(data)
                   if cycles[k] is not None and isinstance(r[j], (int, float))]
            if not ser:
                continue
            first = ser[0][1]
            if 0.05 <= first <= 12:
                caps.append(ser)
            elif 50 <= first <= 200:
                sohs.append(ser)
        for idx, cap in enumerate(caps):
            if idx < len(sohs):
                out.append((cap, sohs[idx]))
    return out


def main():
    wb = openpyxl.load_workbook(M.XLSX, read_only=True, data_only=True, keep_links=False)
    grand_max = 0.0
    for sheet in M.SUMMARY_SHEETS:
        ws = wb[sheet]
        # --- vertical layout: DChg block then SOH block below ---
        blocks = parse_all_blocks(ws)
        pairs = []
        for i in range(len(blocks)):
            if blocks[i][0] == 'dchg':
                soh = next((blocks[k] for k in range(i + 1, len(blocks)) if blocks[k][0] == 'soh'), None)
                if soh:
                    pairs.append((blocks[i], soh))
        sheet_max = 0.0
        n_cmp = 0
        for (_, rate, cap_cells), (_, _, soh_cells) in pairs:
            for cell, cap_series in cap_cells.items():
                if cell not in soh_cells:
                    continue
                ours = soh_from_cap(cap_series)
                for cyc, t in soh_cells[cell]:
                    if cyc in ours and 0 < t < 1000:
                        sheet_max = max(sheet_max, abs(ours[cyc] - t)); n_cmp += 1
        # --- horizontal layout: SOH columns beside capacity columns ---
        for cap, soh in horizontal_pairs(ws):
            ours = soh_from_cap(cap)
            tmap = dict(soh)
            for cyc, t in tmap.items():
                if cyc in ours and 0 < t < 1000:
                    sheet_max = max(sheet_max, abs(ours[cyc] - t)); n_cmp += 1
        grand_max = max(grand_max, sheet_max)
        print(f'  {sheet:9s}: {n_cmp} points compared, max |ΔSOH| = {sheet_max:.4f} %')
    wb.close()
    print(f'\n  GRAND max |Δ(ourSOH − theirSOH)| = {grand_max:.4f} %'
          + ('   ✓ formulas agree' if grand_max < 0.5 else '   ⚠ check'))


if __name__ == '__main__':
    main()

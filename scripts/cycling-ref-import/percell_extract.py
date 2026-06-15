#!/usr/bin/env python3
"""
Robust header-anchored extractor for the messy per-cell sheets.

Facts established in discovery:
  - No explicit cycle column → cycle = data-row index within a block (1-based).
  - Blocks are laid horizontally, gap-separated; each block is a contiguous
    run of metric columns [Chg.Cap, DChg.Cap, DChg.Energy, Med.Volt, (SOH)].
  - A rate label sits in the row above the block's first column.
  - Blocks may also be stacked vertically (multiple header rows) → we scan
    every row for headers, not just the top.
  - Formation blocks usually lack SOH.

extract_cell(ws) → list of blocks:
  { 'rate': str, 'header_row': int, 'cols': {role: idx},
    'series': [ {cycle, chg_cap, dch_cap, dch_energy, med_volt, soh, ce} ] }
"""
import re


def classify(h):
    if not isinstance(h, str):
        return None
    s = h.strip().lower().replace(' ', '')
    if 'soh' in s:
        return 'soh'
    if 'energy' in s or 'энерг' in s:
        return 'dch_energy'
    if 'volt' in s or 'напряж' in s:
        return 'med_volt'
    if 'cap' in s or 'ёмк' in s or 'емк' in s:
        if s.startswith('d'):
            return 'dch_cap'
        if s.startswith('c'):
            return 'chg_cap'
    return None


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def extract_cell(ws):
    rows = list(ws.iter_rows(values_only=True))
    blocks = []
    for hr in range(len(rows)):
        header = rows[hr]
        marks = [(j, classify(c)) for j, c in enumerate(header)]
        marks = [(j, m) for j, m in marks if m]
        if len(marks) < 2:
            continue
        # group adjacent header columns into blocks (a gap >1 col = boundary)
        groups = []
        cur = [marks[0]]
        for (j, m) in marks[1:]:
            if j - cur[-1][0] <= 1:
                cur.append((j, m))
            else:
                groups.append(cur)
                cur = [(j, m)]
        groups.append(cur)

        for g in groups:
            cols = {}
            for (j, m) in g:
                cols.setdefault(m, j)   # first wins if duplicate role
            if 'dch_cap' not in cols:
                continue                # need at least discharge capacity
            c0 = g[0][0]
            # rate label: non-empty strings in the row above, across this block's span
            rate = ''
            if hr - 1 >= 0:
                span = [rows[hr - 1][k] for k in range(c0, g[-1][0] + 1) if k < len(rows[hr - 1])]
                rate = ' '.join(str(x).strip() for x in span if isinstance(x, str) and x.strip())
            # read data until discharge cap goes non-numeric
            series = []
            cyc = 0
            for r in range(hr + 1, len(rows)):
                row = rows[r]
                dch = _num(row[cols['dch_cap']]) if cols['dch_cap'] < len(row) else None
                if dch is None or dch <= 0:
                    # stop at first gap after data started; tolerate a leading blank
                    if series:
                        break
                    else:
                        continue
                cyc += 1
                chg = _num(row[cols['chg_cap']]) if 'chg_cap' in cols and cols['chg_cap'] < len(row) else None
                en = _num(row[cols['dch_energy']]) if 'dch_energy' in cols and cols['dch_energy'] < len(row) else None
                volt = _num(row[cols['med_volt']]) if 'med_volt' in cols and cols['med_volt'] < len(row) else None
                soh = _num(row[cols['soh']]) if 'soh' in cols and cols['soh'] < len(row) else None
                ce = (dch / chg * 100) if (chg and chg > 0) else None
                series.append({'cycle': cyc, 'chg_cap': chg, 'dch_cap': dch,
                               'dch_energy': en, 'med_volt': volt, 'soh': soh, 'ce': ce})
            if series:
                blocks.append({'rate': rate or f'block@{hr + 1}', 'header_row': hr + 1,
                               'cols': cols, 'series': series})
    return blocks


# ── report mode ──
if __name__ == '__main__':
    import os
    import openpyxl
    XLSX = os.environ.get('REF_XLSX') or os.path.expanduser('~/Desktop/файлы коллег/LFP NMC NCA.xlsm')
    SUMMARY = {'LFP-C', 'LFP-LTO', 'LFP 3.0', 'LFP 4.0', 'NMC 3.0', 'NMC-C', 'NCA-C'}
    INDEX = {'Лист1', 'Лист3', 'Список_листов', 'PQ_meta'}
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True, keep_links=False)
    percell = [s for s in wb.sheetnames if s not in SUMMARY and s not in INDEX]
    total_blocks = total_pts = 0
    for s in percell:
        blocks = extract_cell(wb[s])
        total_blocks += len(blocks)
        parts = []
        for b in blocks:
            n = len(b['series'])
            total_pts += n
            has_ce = sum(1 for p in b['series'] if p['ce'] is not None)
            has_soh = sum(1 for p in b['series'] if p['soh'] is not None)
            parts.append(f"«{b['rate'][:22]}»:{n}cyc(ce{has_ce},soh{has_soh})")
        print(f'{s:16s} {len(blocks)} blocks | ' + ' '.join(parts))
    print(f'\nTOTAL: {len(percell)} sheets, {total_blocks} blocks, {total_pts} cycle-points')
    wb.close()

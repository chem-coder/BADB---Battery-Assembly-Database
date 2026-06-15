#!/usr/bin/env python3
"""
Canonical reference import — from the colleague's combined long-format table
`SOH_DSh_cycles` (in ~/Desktop/почта/LFP NMC NCA.xlsm).

This is the clean source: one sheet, labelled columns —
  Название этапа | Тип | Серия | Комментарий | C | n_cycles | n_cycles_complet
  | № | Chg.Cap | DChg.Cap | DChg.Energy | Med.Volt | SOH
— replacing the fragile per-cell parser (which read the same data out of 54
messy sheets). Verified earlier: 0 DChg values differ between the two sources.

Strictness: column indices are resolved by header text (not fixed positions),
blocks are segmented by cycle-number reset, and the script REFUSES to emit SQL
unless its parsed DChg multiset exactly equals the sheet's raw DChg column.

Usage:
  python3 import_from_combined.py            # parse + strict reconcile + plan
  python3 import_from_combined.py --sql OUT  # also write SQL (only if reconciled)
"""
import os
import sys
import collections
import openpyxl

XLSX = os.environ.get('REF_XLSX') or os.path.expanduser('~/Desktop/почта/LFP NMC NCA.xlsm')
SHEET = 'SOH_DSh_cycles'
BATTERY_ID = 1
TAG = 'REF_IMPORT: коллеги LFP NMC NCA (combined)'
PROTO_PREFIXES = ['LFP-C', 'LFP-LTO', 'LFP 3.0', 'LFP 4.0', 'NMC 3.0', 'NMC-C', 'NCA-C']


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def split_series(s):
    s = str(s).strip()
    cell = s.split()[-1] if s.split() else s
    for p in PROTO_PREFIXES:
        if s.startswith(p):
            return p, cell
    return s, cell


def sql_escape(s):
    return str(s).replace("'", "''")


def find_header(rows):
    for i, r in enumerate(rows):
        if any(isinstance(c, str) and 'DChg' in c and 'Cap' in c for c in r):
            return i
    return None


def resolve_cols(hdr):
    def col(pred):
        for j, c in enumerate(hdr):
            if isinstance(c, str) and pred(c.strip()):
                return j
        return None
    ci = {
        'series': col(lambda h: h == 'Серия'),
        'type':   col(lambda h: h == 'Тип'),
        'rate':   col(lambda h: h == 'C'),
        'comment': col(lambda h: h.startswith('Коммент')),
        'cyc':    col(lambda h: h == '№'),
        'chg':    col(lambda h: h.startswith('Chg. Cap')),
        'dch':    col(lambda h: h.startswith('DChg. Cap')),
        'energy': col(lambda h: 'Energy' in h),
        'volt':   col(lambda h: 'Volt' in h),
    }
    return ci


def parse(rows, hr, ci):
    """Segment the long table into blocks (one per cell × rate-step). A new
    block starts when the Серия changes or the cycle № resets (≤ previous)."""
    blocks = []
    cur = None
    last_series = None
    last_cyc = None
    for r in rows[hr + 1:]:
        series = r[ci['series']] if ci['series'] < len(r) else None
        dch = num(r[ci['dch']]) if ci['dch'] < len(r) else None
        rawcyc = r[ci['cyc']] if ci['cyc'] < len(r) else None
        if series is None or dch is None or dch <= 0:
            continue
        if not isinstance(rawcyc, (int, float)):
            continue
        cyc = int(rawcyc)
        new_block = (cur is None) or (series != last_series) or (cyc <= (last_cyc if last_cyc is not None else 0))
        if new_block:
            proto, cell = split_series(series)
            rate = r[ci['rate']] if ci['rate'] is not None and ci['rate'] < len(r) else None
            typ = r[ci['type']] if ci['type'] is not None and ci['type'] < len(r) else None
            label = str(rate).strip() if rate not in (None, '') else 'block'
            if isinstance(typ, str) and 'formation' in typ.lower() and 'formation' not in label.lower():
                label = 'formation ' + label
            cur = {'proto': proto, 'cell': cell, 'rate': label, 'series': []}
            blocks.append(cur)
        chg = num(r[ci['chg']]) if ci['chg'] < len(r) else None
        en = num(r[ci['energy']]) if ci['energy'] is not None and ci['energy'] < len(r) else None
        volt = num(r[ci['volt']]) if ci['volt'] is not None and ci['volt'] < len(r) else None
        ce = (dch / chg * 100) if (chg and chg > 0) else None
        cur['series'].append({'cycle': cyc, 'chg': chg, 'dch': dch, 'energy': en, 'volt': volt, 'ce': ce})
        last_series = series
        last_cyc = cyc
    return blocks


def raw_dchg(rows, hr, ci):
    out = []
    for r in rows[hr + 1:]:
        v = num(r[ci['dch']]) if ci['dch'] < len(r) else None
        if v is not None and v > 0:
            out.append(v)
    return out


def main():
    out_sql = sys.argv[sys.argv.index('--sql') + 1] if '--sql' in sys.argv else None
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True, keep_links=False)
    if SHEET not in wb.sheetnames:
        sys.exit(f'sheet {SHEET!r} not found in {XLSX}')
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()

    hr = find_header(rows)
    if hr is None:
        sys.exit('header row (with "DChg. Cap") not found')
    ci = resolve_cols(rows[hr])
    need = ['series', 'cyc', 'chg', 'dch']
    miss = [k for k in need if ci[k] is None]
    if miss:
        sys.exit(f'required columns missing: {miss}  (resolved: {ci})')
    print(f'header @ row {hr + 1}; columns resolved: '
          + ', '.join(f'{k}={chr(65 + v) if v is not None else "—"}' for k, v in ci.items()))

    blocks = parse(rows, hr, ci)

    # ── STRICT reconciliation: parsed DChg multiset must equal the raw column ──
    parsed = collections.Counter(round(p['dch'], 6) for b in blocks for p in b['series'])
    rawc = collections.Counter(round(v, 6) for v in raw_dchg(rows, hr, ci))
    only_raw = sum((rawc - parsed).values())
    only_parsed = sum((parsed - rawc).values())
    print(f'\nRECONCILE parsed↔sheet DChg: raw={sum(rawc.values())} parsed={sum(parsed.values())} '
          f'| missed={only_raw} extra={only_parsed}')
    if only_raw or only_parsed:
        sys.exit('✗ STRICT FAIL — parsed data does not equal the sheet. No SQL generated.')
    print('✓ exact match — every sheet DChg value is captured, none extra.')

    # ── plan ──
    by_proto = {}
    for b in blocks:
        by_proto.setdefault(b['proto'], []).append(b)
    tot_s = tot_r = 0
    print('\n=== PLAN ===')
    for proto in sorted(by_proto):
        items = by_proto[proto]
        r = sum(len(b['series']) for b in items)
        ce = sum(1 for b in items for p in b['series'] if p['ce'] is not None)
        print(f'  {proto:9s}: {len(items)} sessions, {r} rows, {ce} with CE')
        tot_s += len(items)
        tot_r += r
    print(f'  ──── TOTAL: {tot_s} sessions, {tot_r} rows')

    if not out_sql:
        print('\n(reconciled; pass --sql FILE to write SQL)')
        return

    def n(v):
        return 'NULL' if v is None else f'{v:.6f}'

    def ce_qc(ce):
        return ce if (ce is not None and 50 <= ce <= 105) else None

    lines = ['BEGIN;', "DELETE FROM cycling_sessions WHERE notes LIKE 'REF_IMPORT:%';", '']
    for b in blocks:
        fname = f"{b['proto']} {b['cell']} [{b['rate']}]"
        total = len(b['series'])
        vals = ','.join(
            f"({p['cycle']},{n(p['chg'])},{n(p['dch'])},{n(p['energy'])},{n(p['volt'])},{n(ce_qc(p['ce']))})"
            for p in b['series'])
        lines.append(
            "WITH s AS (\n"
            "  INSERT INTO cycling_sessions (battery_id, equipment_type, file_name, protocol, total_cycles, status, notes)\n"
            f"  VALUES ({BATTERY_ID}, 'generic', '{sql_escape(fname)}', '{sql_escape(b['proto'])}', {total}, 'ready', '{sql_escape(TAG)} | {sql_escape(fname)}')\n"
            "  RETURNING session_id\n)\n"
            "INSERT INTO cycling_cycle_summary\n"
            "  (session_id, cycle_number, charge_capacity_ah, discharge_capacity_ah, discharge_energy_wh, avg_discharge_voltage_v, coulombic_efficiency)\n"
            "SELECT session_id, v.cyc, v.chg::float8, v.dch::float8, v.en::float8, v.volt::float8, v.ce::float8\n"
            f"FROM s, (VALUES {vals}) AS v(cyc, chg, dch, en, volt, ce);")
    lines += ['', 'COMMIT;']
    with open(out_sql, 'w') as f:
        f.write('\n'.join(lines))
    print(f'\nSQL → {out_sql}  ({len(blocks)} sessions)')


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
"""
Phase 1 — discovery. Map the layout of every per-cell sheet so the rich
extractor is built on facts, not guesses. For each sheet we locate the
header cells for each metric (Chg.Cap / DChg.Cap / DChg.Energy / Med.Volt /
SOH), how many blocks (rate-steps) each has, and how the cycle axis works.

Output: one line per sheet (grouped by protocol) + a per-protocol
consistency summary, so we can see whether all cells of a protocol share a
layout (expected) or differ.
"""
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

XLSX = os.environ.get('REF_XLSX') or os.path.expanduser('~/Desktop/файлы коллег/LFP NMC NCA.xlsm')
SUMMARY = {'LFP-C', 'LFP-LTO', 'LFP 3.0', 'LFP 4.0', 'NMC 3.0', 'NMC-C', 'NCA-C'}
INDEX = {'Лист1', 'Лист3', 'Список_листов', 'PQ_meta'}

# Header classifiers (order matters: discharge before charge).
def classify(h):
    if not isinstance(h, str):
        return None
    s = h.strip().lower().replace(' ', '')
    if 'soh' in s:
        return 'soh'
    if 'energy' in s or 'энерг' in s:
        # discharge energy specifically
        return 'dch_energy' if s.startswith('d') else 'chg_energy'
    if 'volt' in s or 'напряж' in s:
        return 'med_volt'
    if 'cap' in s or 'ёмк' in s or 'емк' in s:
        if s.startswith('dch') or s.startswith('dсh') or s.startswith('d'):
            return 'dch_cap'
        if s.startswith('chg') or s.startswith('c'):
            return 'chg_cap'
    return None


def protocol_of(sheet):
    # 'NCA-C A1' -> 'NCA-C'; 'LFP 3.0 n1' -> 'LFP 3.0'; 'LFP-C 2.0 G1' -> 'LFP-C 2.0'
    return re.sub(r'\s+\S+$', '', sheet).strip()


def discover_sheet(ws):
    """Return dict: metric -> list of (row, col) header positions; plus rate labels."""
    found = {}
    rate_labels = []
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), max_col=min(ws.max_column, 60), values_only=True))
    for i, row in enumerate(rows):
        for j, c in enumerate(row):
            m = classify(c)
            if m:
                found.setdefault(m, []).append((i + 1, j + 1))
            elif isinstance(c, str) and re.search(r'\d+\s*C', c) and 'cap' not in c.lower():
                rate_labels.append(c.strip())
    return found, rate_labels


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True, keep_links=False)
    percell = [s for s in wb.sheetnames if s not in SUMMARY and s not in INDEX]
    by_proto = {}
    for s in percell:
        by_proto.setdefault(protocol_of(s), []).append(s)

    print(f'Per-cell sheets: {len(percell)}  across {len(by_proto)} protocols\n')
    sigs = {}
    for proto in sorted(by_proto):
        print(f'━━ {proto} ━━')
        for s in by_proto[proto]:
            ws = wb[s]
            found, rates = discover_sheet(ws)
            sig = tuple(sorted((m, len(v)) for m, v in found.items()))
            sigs.setdefault(proto, set()).add(sig)
            cols = {m: ','.join(f'{get_column_letter(c)}{r}' for r, c in v) for m, v in found.items()}
            print(f'  {s:16s} rows~{ws.max_row:4d} cols~{ws.max_column:3d} | '
                  + ' '.join(f'{m}:{cols[m]}' for m in ['chg_cap','dch_cap','dch_energy','med_volt','soh'] if m in cols)
                  + (f'  rates={rates[:3]}' if rates else ''))
        print()

    print('=== Layout consistency per protocol (1 signature = uniform) ===')
    for proto in sorted(sigs):
        n = len(sigs[proto])
        print(f'  {proto:14s}: {n} distinct layout(s) {"✓ uniform" if n == 1 else "⚠ MIXED"}')
    wb.close()


if __name__ == '__main__':
    main()

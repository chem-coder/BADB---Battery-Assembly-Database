#!/usr/bin/env python3
"""
Reference-data importer for SOH verification (Dima, 2026-06-05).

Reads the colleague workbook 'файлы коллег/LFP NMC NCA.xlsm', parses each
per-protocol summary sheet (one column per cell of DChg.Cap vs cycle), and
emits idempotent SQL that loads each cell as a cycling_session +
cycling_cycle_summary rows.

All imported sessions are TAGGED  notes LIKE 'REF_IMPORT:%'  and attached to
an existing battery, so removal is a single cascading DELETE:
    DELETE FROM cycling_sessions WHERE notes LIKE 'REF_IMPORT:%';

Usage:
    python3 import_ref_cycling.py            # dry-run: print plan only
    python3 import_ref_cycling.py --sql OUT  # also write SQL to OUT
"""
import sys
import os
import openpyxl

XLSX = os.environ.get('REF_XLSX') or os.path.expanduser('~/Desktop/файлы коллег/LFP NMC NCA.xlsm')
SUMMARY_SHEETS = ['LFP-C', 'LFP-LTO', 'LFP 3.0', 'LFP 4.0', 'NMC 3.0', 'NMC-C', 'NCA-C']
BATTERY_ID = 1                       # attach refs to existing battery 1
TAG = 'REF_IMPORT: коллеги LFP NMC NCA'

CAP_LO, CAP_HI = 0.05, 12.0          # plausible Ah discharge-capacity window


def _is_header_cell(c):
    return isinstance(c, str) and ('Cap' in c or 'SOH' in c)


def _is_dchg_header(c):
    """Discharge-capacity header: 'DChg. Cap', 'DCh. Cap' (starts with D, has Cap)."""
    return isinstance(c, str) and 'Cap' in c and c.strip().startswith('D')


def parse_blocks(ws):
    """Multi-block parser. Summary sheets stack one (DChg.Cap block, SOH block)
    pair per rate-step ('0.3C-0.3C', '0.5-1C', ...). Each block restarts cycle
    numbering at 1. Return list of (rate_label, {cell_label: [(cycle, cap)]}).
    One block == one self-normalised cycling phase == one session per cell."""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    # every header row (cap OR soh) — used to bound a block's data range
    header_rows = [i for i, row in enumerate(rows) if any(_is_header_cell(c) for c in row)]
    # discharge-capacity block starts
    dchg = []
    for i, row in enumerate(rows):
        for j, c in enumerate(row):
            if _is_dchg_header(c):
                # rate label sits in the cycle column 1-2 rows above the header
                rate = None
                for up in (i - 1, i - 2):
                    if up >= 0:
                        v = rows[up][j] if j < len(rows[up]) else None
                        if isinstance(v, str) and v.strip() and not _is_header_cell(v):
                            rate = v.strip()
                            break
                dchg.append((rate, i, j))
                break

    result = []
    for rate, hi, cyc_c in dchg:
        hdr = rows[hi]
        end = min([h for h in header_rows if h > hi], default=len(rows))
        data = rows[hi + 1:end]
        cycles = [int(r[cyc_c]) if cyc_c < len(r) and isinstance(r[cyc_c], (int, float)) else None
                  for r in data]
        cells = {}
        for j in range(cyc_c + 1, ws.max_column):
            label = hdr[j] if j < len(hdr) else None
            if label is None or (isinstance(label, str) and 'SOH' in label):
                break
            series = []
            for k, r in enumerate(data):
                v = r[j] if j < len(r) else None
                if cycles[k] is not None and isinstance(v, (int, float)) and CAP_LO <= float(v) <= CAP_HI:
                    series.append((cycles[k], float(v)))
            if series:
                cells[str(label).strip()] = series
        if cells:
            result.append((rate or f'step{len(result) + 1}', cells))
    return result


def sql_escape(s):
    return s.replace("'", "''")


def main():
    out_sql = None
    if '--sql' in sys.argv:
        out_sql = sys.argv[sys.argv.index('--sql') + 1]

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True, keep_links=False)
    plan = []  # (protocol, rate, cell_label, series)
    for sheet in SUMMARY_SHEETS:
        if sheet not in wb.sheetnames:
            print(f'  ! missing sheet {sheet!r}')
            continue
        for rate, cells in parse_blocks(wb[sheet]):
            for label, series in cells.items():
                plan.append((sheet, rate, label, series))
    wb.close()

    # ---- dry-run report ----
    print('\n=== IMPORT PLAN (dry-run) ===')
    by_proto = {}
    for proto, rate, label, series in plan:
        by_proto.setdefault(proto, {}).setdefault(rate, []).append((label, series))
    total_sessions = total_rows = 0
    for proto, steps in by_proto.items():
        print(f'  {proto}:')
        for rate, items in steps.items():
            ncells = len(items)
            nrows = sum(len(s) for _, s in items)
            maxc = max((s[-1][0] for _, s in items), default=0)
            c1 = items[0][1][0][1] if items and items[0][1] else None
            print(f'      step «{rate}»: {ncells} cells × {maxc} cyc, {nrows} rows'
                  + (f'  (c1={c1:.3f}Ah)' if c1 else '') + f'  [{", ".join(l for l,_ in items)}]')
            total_sessions += ncells
            total_rows += nrows
    print(f'  ----  TOTAL: {total_sessions} sessions (cell×step), {total_rows} cycle-summary rows')
    print(f'  battery_id={BATTERY_ID}, status=ready, tag notes={TAG!r}')

    if not out_sql:
        print('\n(dry-run only; pass --sql FILE to also generate SQL)')
        return

    # ---- SQL generation ----
    lines = ['BEGIN;',
             "-- idempotent: drop any prior reference import (cascades summary+datapoints)",
             "DELETE FROM cycling_sessions WHERE notes LIKE 'REF_IMPORT:%';", '']
    for proto, rate, label, series in plan:
        fname = f'{proto} {label} [{rate}]'
        total = series[-1][0]
        # de-dup cycle numbers within a block defensively (keep last)
        seen = {}
        for c, cap in series:
            seen[c] = cap
        values = ','.join(f'({c},{cap:.6f})' for c, cap in sorted(seen.items()))
        lines.append(
            "WITH s AS (\n"
            "  INSERT INTO cycling_sessions (battery_id, equipment_type, file_name, protocol, total_cycles, status, notes)\n"
            f"  VALUES ({BATTERY_ID}, 'generic', '{sql_escape(fname)}', '{sql_escape(proto)}', {total}, 'ready', '{sql_escape(TAG)} | {sql_escape(fname)}')\n"
            "  RETURNING session_id\n"
            ")\n"
            "INSERT INTO cycling_cycle_summary (session_id, cycle_number, discharge_capacity_ah)\n"
            f"SELECT session_id, v.cyc, v.cap FROM s, (VALUES {values}) AS v(cyc, cap);")
    lines.append('')
    lines.append('COMMIT;')
    with open(out_sql, 'w') as f:
        f.write('\n'.join(lines))
    print(f'\nSQL written to {out_sql}  ({len(plan)} INSERT blocks)')


if __name__ == '__main__':
    main()

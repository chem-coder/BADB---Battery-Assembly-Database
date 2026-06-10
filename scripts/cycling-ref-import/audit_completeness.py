#!/usr/bin/env python3
"""
Completeness audit — did the extractor lose any cycles?

For every block we compare:
  extracted = len(series) from extract_cell
  available = count of numeric DChg cells in that block's column, bounded
              by the next header row below (so vertically-stacked blocks
              don't bleed into each other)
A block where available > extracted means a mid-series gap truncated it.
"""
import os
import openpyxl
from percell_extract import extract_cell, classify, _num

XLSX = os.environ.get('REF_XLSX') or os.path.expanduser('~/Desktop/файлы коллег/LFP NMC NCA.xlsm')
SUMMARY = {'LFP-C', 'LFP-LTO', 'LFP 3.0', 'LFP 4.0', 'NMC 3.0', 'NMC-C', 'NCA-C'}
INDEX = {'Лист1', 'Лист3', 'Список_листов', 'PQ_meta'}


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True, keep_links=False)
    percell = [s for s in wb.sheetnames if s not in SUMMARY and s not in INDEX]
    total_extracted = 0
    total_available = 0
    truncated = []
    for sheet in percell:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        blocks = extract_cell(ws)
        header_rows = sorted({b['header_row'] for b in blocks})
        for b in blocks:
            hr = b['header_row']
            dch_col = b['cols']['dch_cap']
            # next header strictly below
            below = [h for h in header_rows if h > hr]
            end = (min(below) - 1) if below else len(rows)
            avail = 0
            for r in range(hr, end):       # rows are 0-based list; hr is 1-based header
                row = rows[r]
                v = _num(row[dch_col]) if dch_col < len(row) else None
                if v is not None and v > 0:
                    avail += 1
            got = len(b['series'])
            total_extracted += got
            total_available += avail
            if avail > got:
                truncated.append((sheet, b['rate'][:25], got, avail))
    wb.close()

    print(f'Blocks extracted total points : {total_extracted}')
    print(f'Raw available DChg points     : {total_available}')
    print(f'Difference                    : {total_available - total_extracted}')
    if truncated:
        print(f'\n⚠ {len(truncated)} blocks possibly TRUNCATED (available > extracted):')
        for s, rate, got, avail in truncated[:30]:
            print(f'   {s:16s} «{rate}»  extracted {got} / available {avail}  (LOST {avail-got})')
    else:
        print('\n✓ No truncation — every numeric DChg point was captured.')


if __name__ == '__main__':
    main()
